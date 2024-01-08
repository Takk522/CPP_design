
from openpyxl import load_workbook
from collections import Counter

excel_file = "sequences.xlsx"

wb = load_workbook(excel_file)
ws = wb.active

# 选择需要统计的列
column_to_count = ws['F']


for i, cell in enumerate(column_to_count, start=1):
    if cell.value:
        total_characters = len(cell.value)
        r_count = Counter(cell.value)['H']

        if total_characters > 0:
            percentage_r = (r_count / total_characters) * 100
            print(f"Percentage of 'R' in cell {i}: {percentage_r:.2f}%")

            ws.cell(row=i, column=7, value=percentage_r)
        else:
            print(f"No characters found in cell {i}.")

wb.save(excel_file)
wb.close()


'''

import pandas as pd

# 读取Excel文件
excel_file = "sequences.xlsx"
df = pd.read_excel(excel_file)

# 删除第二列中的重复字符串及其对应的行
df_no_duplicates = df.drop_duplicates(subset=[df.columns[1]])

# 保存处理后的数据到新的Excel文件
output_file = "sequences_no_duplicates.xlsx"
df_no_duplicates.to_excel(output_file, index=False)

print(f"Duplicate rows removed. Result saved to {output_file}")
'''